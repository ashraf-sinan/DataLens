import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
import re


class ExcelExporter:
    """Export analysis results to Excel with visualizations and hyperlinks."""

    def __init__(self, results: Dict, group_column: str = None):
        self.results = results
        self.group_column = group_column
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.viz_charts = []  # Store charts to add to visualization sheet

    def create_safe_sheet_name(self, name: str) -> str:
        """Create a safe sheet name (max 31 chars, no invalid characters)."""
        # Remove invalid characters
        safe_name = re.sub(r'[\[\]:*?/\\]', '_', str(name))
        # Truncate to 31 characters
        if len(safe_name) > 31:
            safe_name = safe_name[:31]
        return safe_name

    def export_ungrouped(self, output_path: str):
        """Export ungrouped analysis results."""
        # Create index sheet
        index_ws = self.wb.create_sheet("Index", 0)
        self._format_index_sheet(index_ws)

        # Create visualization overview sheet
        viz_ws = self.wb.create_sheet("Visualizations", 1)

        row_idx = 2

        # Create a sheet for each column
        for idx, (column_name, column_data) in enumerate(self.results.items(), start=2):
            sheet_name = self.create_safe_sheet_name(column_name)
            ws = self.wb.create_sheet(sheet_name, idx)

            # Add to index with hyperlink
            index_ws.cell(row=row_idx, column=1, value=row_idx - 1)
            index_ws.cell(row=row_idx, column=2, value=column_name)
            index_ws.cell(row=row_idx, column=3, value=column_data['type'].upper())

            # Create hyperlink
            index_ws.cell(row=row_idx, column=4, value="Go to Sheet")
            index_ws.cell(row=row_idx, column=4).hyperlink = f"#{sheet_name}!A1"
            index_ws.cell(row=row_idx, column=4).font = Font(color="0563C1", underline="single")

            row_idx += 1

            # Fill the column sheet and collect charts for visualization sheet
            if column_data['type'] == 'quantitative':
                self._create_quantitative_sheet(ws, column_name, column_data['data'], sheet_name)
            else:
                self._create_qualitative_sheet(ws, column_name, column_data['data'])

        # Now add all collected charts to visualization sheet
        self._populate_viz_overview(viz_ws)

        # Save workbook
        self.wb.save(output_path)

    def export_grouped(self, output_path: str):
        """Export grouped analysis results."""
        # Create index sheet
        index_ws = self.wb.create_sheet("Index", 0)
        self._format_index_sheet_grouped(index_ws)

        # Create visualization overview sheet
        viz_ws = self.wb.create_sheet("Visualizations", 1)

        row_idx = 2
        sheet_idx = 2

        # Create sheets for each group
        for group_name, group_data in self.results.items():
            # Add group header to index
            index_ws.cell(row=row_idx, column=1, value=f"GROUP: {group_name}")
            index_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            row_idx += 1

            # Process each column in the group
            for column_name, column_data in group_data['columns'].items():
                sheet_name = self.create_safe_sheet_name(f"{group_name}_{column_name}")
                ws = self.wb.create_sheet(sheet_name, sheet_idx)
                sheet_idx += 1

                # Add to index
                index_ws.cell(row=row_idx, column=1, value=column_name)
                index_ws.cell(row=row_idx, column=2, value=column_data['type'].upper())
                index_ws.cell(row=row_idx, column=3, value="Go to Sheet")
                index_ws.cell(row=row_idx, column=3).hyperlink = f"#{sheet_name}!A1"
                index_ws.cell(row=row_idx, column=3).font = Font(color="0563C1", underline="single")

                row_idx += 1

                # Fill the sheet
                if column_data['type'] == 'quantitative':
                    self._create_quantitative_sheet(ws, f"{group_name} - {column_name}", column_data['data'], sheet_name)
                else:
                    self._create_qualitative_sheet(ws, f"{group_name} - {column_name}", column_data['data'])

            row_idx += 1  # Spacing between groups

        # Now add all collected charts to visualization sheet
        self._populate_viz_overview(viz_ws)

        # Save workbook
        self.wb.save(output_path)

    def _format_index_sheet(self, ws):
        """Format the index sheet."""
        ws.title = "Index"
        ws['A1'] = "Excel Analysis Results - Column Index"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')

        # Headers
        headers = ["#", "Column Name", "Type", "Link"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

    def _format_index_sheet_grouped(self, ws):
        """Format the index sheet for grouped results."""
        ws.title = "Index"
        ws['A1'] = f"Excel Analysis Results - Grouped by {self.group_column}"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:C1')

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _populate_viz_overview(self, ws):
        """Populate the visualization overview sheet with charts."""
        ws.title = "Visualizations"
        ws['A1'] = "Visualizations Overview"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:P1')

        ws['A3'] = "This sheet contains visualizations for all quantitative columns."
        ws['A3'].font = Font(size=11)

        # Add all collected charts to this sheet
        chart_row = 5
        chart_col_positions = ['A', 'I', 'Q']  # Three columns of charts
        chart_idx = 0

        for chart_info in self.viz_charts:
            col_pos = chart_col_positions[chart_idx % 3]

            # Add section label
            label_cell = ws[f'{col_pos}{chart_row}']
            label_cell.value = f"{chart_info['column_name']} - {chart_info['chart_type']}"
            label_cell.font = Font(bold=True, size=10)

            # Add the chart
            ws.add_chart(chart_info['chart'], f'{col_pos}{chart_row + 1}')

            chart_idx += 1

            # Move to next row after every 3 charts
            if chart_idx % 3 == 0:
                chart_row += 18  # Space for chart height

        ws.column_dimensions['A'].width = 2

    def _create_quantitative_sheet(self, ws, column_name: str, data: Dict[str, Any], sheet_name: str = None):
        """Create sheet for quantitative column analysis with multiple visualizations."""
        # Title
        ws['A1'] = f"Analysis: {column_name}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:E1')

        # Statistics section
        ws['A3'] = "Statistical Summary"
        ws['A3'].font = Font(bold=True, size=12)

        stats = [
            ("Count", data['count']),
            ("Minimum", f"{data['min']:.2f}"),
            ("25th Percentile", f"{data['percentile_25']:.2f}"),
            ("Median (50th)", f"{data['percentile_50']:.2f}"),
            ("75th Percentile", f"{data['percentile_75']:.2f}"),
            ("Maximum", f"{data['max']:.2f}"),
            ("Average", f"{data['average']:.2f}"),
            ("Sum", f"{data['sum']:.2f}"),
            ("% of Total", f"{data['percent_of_total']:.2f}%"),
        ]

        row = 4
        for stat_name, stat_value in stats:
            ws.cell(row=row, column=1, value=stat_name).font = Font(bold=True)
            ws.cell(row=row, column=2, value=stat_value)
            row += 1

        # Create box plot data for visualization
        box_data_row = row + 1
        ws.cell(row=box_data_row, column=7, value="Box Plot Data").font = Font(bold=True)
        ws.cell(row=box_data_row + 1, column=7, value="Min")
        ws.cell(row=box_data_row + 1, column=8, value=data['min'])
        ws.cell(row=box_data_row + 2, column=7, value="Q1")
        ws.cell(row=box_data_row + 2, column=8, value=data['percentile_25'])
        ws.cell(row=box_data_row + 3, column=7, value="Median")
        ws.cell(row=box_data_row + 3, column=8, value=data['percentile_50'])
        ws.cell(row=box_data_row + 4, column=7, value="Q3")
        ws.cell(row=box_data_row + 4, column=8, value=data['percentile_75'])
        ws.cell(row=box_data_row + 5, column=7, value="Max")
        ws.cell(row=box_data_row + 5, column=8, value=data['max'])

        # Frequency distribution section
        freq_start_row = row + 2
        ws.cell(row=freq_start_row, column=1, value="Frequency Distribution").font = Font(bold=True, size=12)

        freq_headers = ["Value", "Frequency", "% of Count", "Value Sum", "% of Total"]
        freq_header_row = freq_start_row + 1
        for col, header in enumerate(freq_headers, start=1):
            cell = ws.cell(row=freq_header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Frequency data
        freq_data_start = freq_header_row + 1
        for idx, freq_item in enumerate(data['frequency']):
            row = freq_data_start + idx
            ws.cell(row=row, column=1, value=freq_item['value'])
            ws.cell(row=row, column=2, value=freq_item['frequency'])
            ws.cell(row=row, column=3, value=f"{freq_item['percentage']:.2f}%")
            ws.cell(row=row, column=4, value=f"{freq_item['value_sum']:.2f}")
            ws.cell(row=row, column=5, value=f"{freq_item['percent_of_total_column']:.2f}%")

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Create visualizations
        charts_created = []

        # 1. Create histogram (bar chart for frequency distribution)
        if len(data['frequency']) > 0 and len(data['frequency']) <= 50:
            histogram = BarChart()
            histogram.type = "col"
            histogram.title = f"Histogram - Frequency Distribution"
            histogram.y_axis.title = 'Frequency'
            histogram.x_axis.title = 'Value'

            data_ref = Reference(ws, min_col=2, min_row=freq_header_row, max_row=freq_data_start + len(data['frequency']) - 1)
            cats = Reference(ws, min_col=1, min_row=freq_header_row + 1, max_row=freq_data_start + len(data['frequency']) - 1)

            histogram.add_data(data_ref, titles_from_data=True)
            histogram.set_categories(cats)
            histogram.height = 10
            histogram.width = 15

            # Place histogram on the sheet
            ws.add_chart(histogram, f"G{freq_start_row}")

            # Create a copy for visualization sheet
            histogram_viz = BarChart()
            histogram_viz.type = "col"
            histogram_viz.title = f"{column_name} - Histogram"
            histogram_viz.y_axis.title = 'Frequency'
            histogram_viz.x_axis.title = 'Value'

            if sheet_name:
                data_ref_viz = Reference(ws, min_col=2, min_row=freq_header_row, max_row=freq_data_start + len(data['frequency']) - 1)
                cats_viz = Reference(ws, min_col=1, min_row=freq_header_row + 1, max_row=freq_data_start + len(data['frequency']) - 1)
                histogram_viz.add_data(data_ref_viz, titles_from_data=True)
                histogram_viz.set_categories(cats_viz)
                histogram_viz.height = 8
                histogram_viz.width = 12

                self.viz_charts.append({
                    'column_name': column_name,
                    'chart_type': 'Histogram',
                    'chart': histogram_viz
                })

        # 2. Create distribution density diagram (line chart showing value distribution)
        if len(data['frequency']) > 0 and len(data['frequency']) <= 50:
            density_chart = LineChart()
            density_chart.title = f"Distribution Density"
            density_chart.y_axis.title = 'Frequency'
            density_chart.x_axis.title = 'Value'
            density_chart.style = 13

            density_data_ref = Reference(ws, min_col=2, min_row=freq_header_row, max_row=freq_data_start + len(data['frequency']) - 1)
            density_cats = Reference(ws, min_col=1, min_row=freq_header_row + 1, max_row=freq_data_start + len(data['frequency']) - 1)

            density_chart.add_data(density_data_ref, titles_from_data=True)
            density_chart.set_categories(density_cats)
            density_chart.height = 10
            density_chart.width = 15

            # Place density chart on sheet
            ws.add_chart(density_chart, f"V{freq_start_row}")

            # Create copy for visualization sheet
            density_chart_viz = LineChart()
            density_chart_viz.title = f"{column_name} - Distribution Density"
            density_chart_viz.y_axis.title = 'Frequency'
            density_chart_viz.x_axis.title = 'Value'
            density_chart_viz.style = 13

            if sheet_name:
                density_data_ref_viz = Reference(ws, min_col=2, min_row=freq_header_row, max_row=freq_data_start + len(data['frequency']) - 1)
                density_cats_viz = Reference(ws, min_col=1, min_row=freq_header_row + 1, max_row=freq_data_start + len(data['frequency']) - 1)
                density_chart_viz.add_data(density_data_ref_viz, titles_from_data=True)
                density_chart_viz.set_categories(density_cats_viz)
                density_chart_viz.height = 8
                density_chart_viz.width = 12

                self.viz_charts.append({
                    'column_name': column_name,
                    'chart_type': 'Distribution Density',
                    'chart': density_chart_viz
                })

    def _create_qualitative_sheet(self, ws, column_name: str, data: list):
        """Create sheet for qualitative column analysis."""
        # Title
        ws['A1'] = f"Analysis: {column_name}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')

        # Data section
        ws['A3'] = "Category Distribution"
        ws['A3'].font = Font(bold=True, size=12)

        headers = ["Label", "Frequency", "Percentage"]
        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Data rows
        data_start = header_row + 1
        for idx, item in enumerate(data):
            row = data_start + idx
            ws.cell(row=row, column=1, value=item['label'])
            ws.cell(row=row, column=2, value=item['frequency'])
            ws.cell(row=row, column=3, value=f"{item['percentage']:.2f}%")

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Create pie chart
        if len(data) > 0 and len(data) <= 20:
            chart = PieChart()
            chart.title = f"{column_name} - Distribution"

            labels = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + len(data) - 1)
            data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=data_start + len(data) - 1)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 12
            chart.width = 20

            ws.add_chart(chart, "E4")

        # Create bar chart for many categories
        elif len(data) > 20:
            chart = BarChart()
            chart.type = "col"
            chart.title = f"{column_name} - Distribution"
            chart.y_axis.title = 'Frequency'
            chart.x_axis.title = 'Category'

            data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=data_start + len(data) - 1)
            cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + len(data) - 1)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 15
            chart.width = 25

            ws.add_chart(chart, "E4")
